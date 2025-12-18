import re
import io
import csv
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.database import get_db
from app.models.core import Model, Series, Manufacturer, EquipmentType
from app.models.templates import AmazonProductType, ProductTypeField, EquipmentTypeProductType

router = APIRouter(prefix="/export", tags=["export"])

class ExportPreviewRequest(BaseModel):
    model_ids: List[int]
    listing_type: str = "individual"  # "individual" or "parent_child"

class ExportRowData(BaseModel):
    model_id: int
    model_name: str
    data: List[str | None]

class ExportPreviewResponse(BaseModel):
    headers: List[List[str | None]]
    rows: List[ExportRowData]
    template_code: str

@router.post("/preview", response_model=ExportPreviewResponse)
def generate_export_preview(request: ExportPreviewRequest, db: Session = Depends(get_db)):
    if not request.model_ids:
        raise HTTPException(status_code=400, detail="No models selected")
    
    models = db.query(Model).filter(Model.id.in_(request.model_ids)).all()
    if not models:
        raise HTTPException(status_code=404, detail="No models found")
    
    equipment_type_ids = set(m.equipment_type_id for m in models)
    if len(equipment_type_ids) > 1:
        raise HTTPException(
            status_code=400, 
            detail="All selected models must have the same equipment type for export"
        )
    
    equipment_type_id = list(equipment_type_ids)[0]
    
    link = db.query(EquipmentTypeProductType).filter(
        EquipmentTypeProductType.equipment_type_id == equipment_type_id
    ).first()
    
    if not link:
        equipment_type = db.query(EquipmentType).filter(EquipmentType.id == equipment_type_id).first()
        raise HTTPException(
            status_code=400, 
            detail=f"No Amazon template linked to equipment type: {equipment_type.name if equipment_type else 'Unknown'}"
        )
    
    product_type = db.query(AmazonProductType).filter(
        AmazonProductType.id == link.product_type_id
    ).first()
    
    if not product_type:
        raise HTTPException(status_code=404, detail="Template not found")
    
    fields = db.query(ProductTypeField).filter(
        ProductTypeField.product_type_id == product_type.id
    ).order_by(ProductTypeField.order_index).all()
    
    header_rows = product_type.header_rows or []
    
    equipment_type = db.query(EquipmentType).filter(EquipmentType.id == equipment_type_id).first()
    
    rows = []
    for model in models:
        series = db.query(Series).filter(Series.id == model.series_id).first()
        manufacturer = db.query(Manufacturer).filter(Manufacturer.id == series.manufacturer_id).first() if series else None
        
        row_data: List[str | None] = []
        for field in fields:
            value = get_field_value(field, model, series, manufacturer, equipment_type, request.listing_type)
            row_data.append(value)
        
        rows.append(ExportRowData(
            model_id=model.id,
            model_name=model.name,
            data=row_data
        ))
    
    return ExportPreviewResponse(
        headers=header_rows,
        rows=rows,
        template_code=product_type.code
    )


from datetime import datetime

def build_export_data(request: ExportPreviewRequest, db: Session):
    """Build export data (headers and rows) for the given models."""
    if not request.model_ids:
        raise HTTPException(status_code=400, detail="No models selected")
    
    models = db.query(Model).filter(Model.id.in_(request.model_ids)).all()
    if not models:
        raise HTTPException(status_code=404, detail="No models found")
    
    equipment_type_ids = set(m.equipment_type_id for m in models)
    if len(equipment_type_ids) > 1:
        raise HTTPException(
            status_code=400, 
            detail="All selected models must have the same equipment type for export"
        )
    
    equipment_type_id = list(equipment_type_ids)[0]
    
    link = db.query(EquipmentTypeProductType).filter(
        EquipmentTypeProductType.equipment_type_id == equipment_type_id
    ).first()
    
    if not link:
        equipment_type = db.query(EquipmentType).filter(EquipmentType.id == equipment_type_id).first()
        raise HTTPException(
            status_code=400, 
            detail=f"No Amazon template linked to equipment type: {equipment_type.name if equipment_type else 'Unknown'}"
        )
    
    product_type = db.query(AmazonProductType).filter(
        AmazonProductType.id == link.product_type_id
    ).first()
    
    if not product_type:
        raise HTTPException(status_code=404, detail="Template not found")
    
    fields = db.query(ProductTypeField).filter(
        ProductTypeField.product_type_id == product_type.id
    ).order_by(ProductTypeField.order_index).all()
    
    header_rows = product_type.header_rows or []
    
    equipment_type = db.query(EquipmentType).filter(EquipmentType.id == equipment_type_id).first()
    
    first_model = models[0]
    first_series = db.query(Series).filter(Series.id == first_model.series_id).first()
    first_manufacturer = db.query(Manufacturer).filter(Manufacturer.id == first_series.manufacturer_id).first() if first_series else None
    
    mfr_name = normalize_for_url(first_manufacturer.name) if first_manufacturer else 'Unknown'
    series_name = normalize_for_url(first_series.name) if first_series else 'Unknown'
    date_str = datetime.now().strftime('%Y%m%d')
    filename_base = f"Amazon_{mfr_name}_{series_name}_{date_str}"
    
    data_rows = []
    for model in models:
        series = db.query(Series).filter(Series.id == model.series_id).first()
        manufacturer = db.query(Manufacturer).filter(Manufacturer.id == series.manufacturer_id).first() if series else None
        
        row_data = []
        for field in fields:
            value = get_field_value(field, model, series, manufacturer, equipment_type, request.listing_type)
            row_data.append(value if value else '')
        
        data_rows.append(row_data)
    
    return header_rows, data_rows, filename_base


@router.post("/download/xlsx")
def download_xlsx(request: ExportPreviewRequest, db: Session = Depends(get_db)):
    """Download export as XLSX file."""
    header_rows, data_rows, filename_base = build_export_data(request, db)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    
    row_styles = [
        (Font(bold=True, color="FFFFFF"), PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")),
        (Font(color="FFFFFF"), PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")),
        (Font(bold=True, color="FFFFFF"), PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")),
        (Font(bold=True), PatternFill(start_color="8BC34A", end_color="8BC34A", fill_type="solid")),
        (Font(size=9), PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")),
        (Font(italic=True, size=9), PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")),
    ]
    
    current_row = 1
    for row_idx, header_row in enumerate(header_rows):
        for col_idx, value in enumerate(header_row):
            cell = ws.cell(row=current_row, column=col_idx + 1, value=value or '')
            if row_idx < len(row_styles):
                cell.font = row_styles[row_idx][0]
                cell.fill = row_styles[row_idx][1]
            cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1
    
    for data_row in data_rows:
        for col_idx, value in enumerate(data_row):
            ws.cell(row=current_row, column=col_idx + 1, value=value or '')
        current_row += 1
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = min(len(str(cell.value)), 50)
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{filename_base}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/download/xlsm")
def download_xlsm(request: ExportPreviewRequest, db: Session = Depends(get_db)):
    """Download export as XLSM file (macro-enabled workbook)."""
    header_rows, data_rows, filename_base = build_export_data(request, db)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    
    row_styles = [
        (Font(bold=True, color="FFFFFF"), PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")),
        (Font(color="FFFFFF"), PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")),
        (Font(bold=True, color="FFFFFF"), PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")),
        (Font(bold=True), PatternFill(start_color="8BC34A", end_color="8BC34A", fill_type="solid")),
        (Font(size=9), PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")),
        (Font(italic=True, size=9), PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")),
    ]
    
    current_row = 1
    for row_idx, header_row in enumerate(header_rows):
        for col_idx, value in enumerate(header_row):
            cell = ws.cell(row=current_row, column=col_idx + 1, value=value or '')
            if row_idx < len(row_styles):
                cell.font = row_styles[row_idx][0]
                cell.fill = row_styles[row_idx][1]
            cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1
    
    for data_row in data_rows:
        for col_idx, value in enumerate(data_row):
            ws.cell(row=current_row, column=col_idx + 1, value=value or '')
        current_row += 1
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = min(len(str(cell.value)), 50)
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{filename_base}.xlsm"
    return StreamingResponse(
        output,
        media_type="application/vnd.ms-excel.sheet.macroEnabled.12",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/download/csv")
def download_csv(request: ExportPreviewRequest, db: Session = Depends(get_db)):
    """Download export as CSV file."""
    header_rows, data_rows, filename_base = build_export_data(request, db)
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    for header_row in header_rows:
        writer.writerow([v or '' for v in header_row])
    
    for data_row in data_rows:
        writer.writerow([v or '' for v in data_row])
    
    content = output.getvalue().encode('utf-8')
    
    filename = f"{filename_base}.csv"
    return StreamingResponse(
        iter([content]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


IMAGE_FIELD_TO_NUMBER = {
    'main_product_image_locator': '001',
    'other_product_image_locator_1': '002',
    'other_product_image_locator_2': '003',
    'other_product_image_locator_3': '004',
    'other_product_image_locator_4': '005',
    'other_product_image_locator_5': '006',
    'other_product_image_locator_6': '007',
    'other_product_image_locator_7': '008',
    'other_product_image_locator_8': '009',
    'swatch_product_image_locator': '010',
}

def normalize_for_url(name: str) -> str:
    """Normalize a name for use in URL paths/filenames.
    Removes spaces, special characters, and non-alphanumeric characters.
    Example: "Fender USA" -> "FenderUSA", "Tone-Master" -> "ToneMaster"
    """
    if not name:
        return ''
    result = re.sub(r'[^a-zA-Z0-9]', '', name)
    return result

def substitute_placeholders(value: str, model: Model, series, manufacturer, equipment_type, is_image_url: bool = False) -> str:
    result = value
    mfr_name = manufacturer.name if manufacturer else ''
    series_name = series.name if series else ''
    model_name = model.name if model else ''
    equip_type = equipment_type.name if equipment_type else ''
    
    if is_image_url:
        mfr_name_norm = normalize_for_url(mfr_name)
        series_name_norm = normalize_for_url(series_name)
        model_name_norm = normalize_for_url(model_name)
        
        result = result.replace('[Manufacturer_Name]', mfr_name_norm)
        result = result.replace('[Series_Name]', series_name_norm)
        result = result.replace('[Model_Name]', model_name_norm)
        result = result.replace('[MANUFACTURER_NAME]', mfr_name_norm)
        result = result.replace('[SERIES_NAME]', series_name_norm)
        result = result.replace('[MODEL_NAME]', model_name_norm)
    else:
        result = result.replace('[MANUFACTURER_NAME]', mfr_name)
        result = result.replace('[SERIES_NAME]', series_name)
        result = result.replace('[MODEL_NAME]', model_name)
        result = result.replace('[Manufacturer_Name]', mfr_name)
        result = result.replace('[Series_Name]', series_name)
        result = result.replace('[Model_Name]', model_name)
    
    result = result.replace('[EQUIPMENT_TYPE]', equip_type)
    result = result.replace('[Equipment_Type]', equip_type)
    
    return result

def get_image_field_key(field_name: str) -> str | None:
    """Extract the base image field key from a full Amazon field name.
    Returns the key if it matches a known product image field, None otherwise.
    """
    for key in IMAGE_FIELD_TO_NUMBER.keys():
        if field_name.startswith(key):
            return key
    return None

def is_image_url_field(field_name: str) -> bool:
    """Check if a field is a product image URL field that needs special processing."""
    return get_image_field_key(field_name) is not None

def get_field_value(field: ProductTypeField, model: Model, series, manufacturer, equipment_type=None, listing_type: str = "individual") -> str | None:
    field_name_lower = field.field_name.lower()
    is_image_field = is_image_url_field(field.field_name)
    
    if 'contribution_sku' in field_name_lower and listing_type == 'individual':
        return model.parent_sku if model.parent_sku else None
    
    if field.custom_value:
        return substitute_placeholders(field.custom_value, model, series, manufacturer, equipment_type, is_image_url=is_image_field)
    
    if field.selected_value:
        return field.selected_value
    
    if 'item_name' in field_name_lower or 'product_name' in field_name_lower or 'title' in field_name_lower:
        mfr_name = manufacturer.name if manufacturer else ''
        series_name = series.name if series else ''
        return f"{mfr_name} {series_name} {model.name} Cover"
    
    if 'brand' in field_name_lower or 'brand_name' in field_name_lower:
        return manufacturer.name if manufacturer else None
    
    if 'model' in field_name_lower or 'model_number' in field_name_lower or 'model_name' in field_name_lower:
        return model.name
    
    if 'manufacturer' in field_name_lower:
        return manufacturer.name if manufacturer else None
    
    return None
